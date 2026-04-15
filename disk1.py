import boto3
import time
from openpyxl import Workbook
from openpyxl.styles import Font

ssm = boto3.client('ssm')
ec2 = boto3.client('ec2')

instances = ec2.describe_instances()

instance_data = []

for res in instances['Reservations']:
    for inst in res['Instances']:
        if inst['State']['Name'] == 'running':

            instance_id = inst['InstanceId']

            name = "N/A"
            for tag in inst.get('Tags', []):
                if tag['Key'] == 'Name':
                    name = tag['Value']

            public_ip = inst.get('PublicIpAddress', 'N/A')
            private_ip = inst.get('PrivateIpAddress', 'N/A')  # 🔥 NEW

            instance_data.append({
                "id": instance_id,
                "name": name,
                "public_ip": public_ip,
                "private_ip": private_ip
            })

wb = Workbook()
ws = wb.active
ws.title = "Disk_Report"

row = 1

for inst in instance_data:
    instance_id = inst["id"]

    try:
        response = ssm.send_command(
            InstanceIds=[instance_id],
            DocumentName="AWS-RunShellScript",
            Parameters={'commands': [
                'df -TH',
                'cat /etc/os-release | grep PRETTY_NAME'
            ]}
        )

        command_id = response['Command']['CommandId']
        time.sleep(5)

        output = ssm.get_command_invocation(
            CommandId=command_id,
            InstanceId=instance_id
        )

        result = output['StandardOutputContent']
        lines = result.split("\n")

        # 🔥 OS clean
        os_version = "N/A"
        for l in lines:
            if "PRETTY_NAME" in l:
                os_version = l.split("=")[1].replace('"', '')

        # 🔥 Instance details
        ws.cell(row=row, column=1, value="Instance ID").font = Font(bold=True)
        ws.cell(row=row, column=2, value=inst["id"])
        row += 1

        ws.cell(row=row, column=1, value="Instance Name").font = Font(bold=True)
        ws.cell(row=row, column=2, value=inst["name"])
        row += 1

        ws.cell(row=row, column=1, value="Public IP").font = Font(bold=True)
        ws.cell(row=row, column=2, value=inst["public_ip"])
        row += 1

        ws.cell(row=row, column=1, value="Private IP").font = Font(bold=True)  # 🔥 NEW
        ws.cell(row=row, column=2, value=inst["private_ip"])
        row += 1

        ws.cell(row=row, column=1, value="OS Version").font = Font(bold=True)
        ws.cell(row=row, column=2, value=os_version)
        row += 1

        # 🔥 Table header
        headers = ["Filesystem", "Type", "Size", "Used", "Avail", "Use%", "Mounted_on"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = Font(bold=True)

        row += 1

        # 🔥 Table data
        for line in lines:
            if line.strip() == "" or "Filesystem" in line or "PRETTY_NAME" in line:
                continue

            cols = line.split()
            for col_idx, val in enumerate(cols, 1):
                ws.cell(row=row, column=col_idx, value=val)

            row += 1

        row += 2

    except Exception as e:
        print(f"Skipping {instance_id}")

wb.save("disk_report.xlsx")

print("✅ Report with Private IP created")